import openpyxl as xl

accountSummaryHeading = ['Account Number', 'Account Type', 'Bank Name', 'IFSC', 'Status']
profileHeading = ['Name', 'Date of Birth', 'Mobile', 'Address', 'Email', 'PAN']
transactionHeading = ['Type', 'Account', 'Transaction Date', 'Narration']
files = 'DataStore.xlsx'


def createDataStore(accountSummaryHeading=None, profileHeading=None, transactionHeading=None):
    # check if file already exists
    try:
        wb = xl.load_workbook(files)
        ws = wb.worksheets[0]  # select first worksheet
    except FileNotFoundError:
        # create a workbook
        workbook = xl.Workbook()
        # create different sheets to store the data
        accountSummarySheet = workbook.active
        accountSummarySheet.title = 'Account Summary'
        profileSheet = workbook.create_sheet(title='Profiles')
        transactionSheet = workbook.create_sheet(title='Transactions')
        # write data heading rows
        accountSummarySheet.append(accountSummaryHeading)
        profileSheet.append(profileHeading)
        transactionSheet.append(transactionHeading)
        workbook.save(files)

def addValuesToExcel(values, sheetName):
    # open created workbook
    workbook = xl.load_workbook(files)
    # open required worksheet
    sheet = workbook[sheetName]
    # append values to worksheet
    sheet.append(list(values.values()))
    workbook.save(files)

if __name__ == '__main__':
    createDataStore(accountSummaryHeading, profileHeading, transactionHeading)
